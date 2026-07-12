import logging
import os

import pandas as pd
from openpyxl import load_workbook

from .jobs import finished_jobs, progress_queues

logger = logging.getLogger(__name__)

COLUNAS_FINAIS = [
    "CPF", "MASP", "NOME", "COD_C", "N_CONTRATO", "NOME_CONV",
    "VALOR_LANCADO", "VALOR_DESCONTADO", "DIFERENCA", "DATA", "DATA_LANC",
]


def _formatar_data(serie):
    return (
        serie.astype(str)
        .str.replace(r"\D", "", regex=True)
        .str.zfill(8)
        .apply(lambda x: f"{x[:2]}/{x[2:4]}/{x[4:]}")
    )


def _escrever_aba(wb, nome_aba, df):
    ws = wb.create_sheet(title=nome_aba)
    ws.append(list(df.columns))
    for i in range(len(df)):
        ws.append(df.iloc[i].tolist())
    for row in ws.iter_rows(min_row=2, min_col=7, max_col=9):
        for cell in row:
            cell.number_format = "R$ #,##0.00"
    return ws


def _formatar_aba_existente(ws):
    for row in ws.iter_rows(min_row=2, min_col=7, max_col=9):
        for cell in row:
            cell.number_format = "R$ #,##0.00"


def _ajustar_largura(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 5


def converter_txt_para_xlsx(caminho_txt, saida, job_id):
    """
    Lê um arquivo TXT delimitado por ';' com o layout esperado de folha de
    pagamento/consignado e gera uma planilha Excel com 3 abas:
    ABA_1 (todos os registros), DESCONTADOS e NAO_DESCONTADOS.

    Publica progresso na fila associada ao job_id (ver app/jobs.py) para que
    a rota /progresso/<job_id> transmita o andamento via Server-Sent Events.
    """
    q = progress_queues[job_id]

    def emit(msg, pct):
        q.put({"msg": msg, "pct": pct})

    try:
        emit("Lendo arquivo .txt...", 5)
        df = pd.read_csv(caminho_txt, sep=";", header=None, dtype=str, encoding="latin1")
        df = df.map(lambda x: x.strip() if isinstance(x, str) else x)

        if df.shape[1] < 18:
            raise ValueError(
                "O arquivo não tem o número de colunas esperado. "
                "Verifique se o layout do TXT está correto."
            )

        emit("Parseando colunas e formatando dados...", 20)
        df_base = df[[1, 2, 3, 7, 8, 10, 11, 12, 14, 17]].copy()
        df_base["N_CONTRATO"] = df[9]
        df_base.columns = [
            "CPF", "MASP", "NOME", "COD_C", "NOME_CONV",
            "VALOR", "VALOR_DESC_ORIG", "DIF", "DATA", "DATA_LANC", "N_CONTRATO",
        ]
        df_base = df_base[[
            "CPF", "MASP", "NOME", "COD_C", "N_CONTRATO", "NOME_CONV",
            "VALOR", "VALOR_DESC_ORIG", "DIF", "DATA", "DATA_LANC",
        ]]
        df_base["COD_C"] = df_base["COD_C"].astype(str).str.extract(r"(\d+)")[0].str.zfill(3)
        df_base["N_CONTRATO"] = df_base["N_CONTRATO"].str.strip().str.lstrip("0")

        emit("Calculando valores financeiros...", 40)
        df_base["VALOR_LANCADO"] = pd.to_numeric(df_base["VALOR"], errors="coerce") / 100
        df_base["VALOR_DESCONTADO"] = pd.to_numeric(df_base["VALOR_DESC_ORIG"], errors="coerce") / 100
        df_base["VALOR_LANCADO"] = df_base["VALOR_LANCADO"].round(2).fillna(0)
        df_base["VALOR_DESCONTADO"] = df_base["VALOR_DESCONTADO"].round(2).fillna(0)
        df_base["DIFERENCA"] = (df_base["VALOR_LANCADO"] - df_base["VALOR_DESCONTADO"]).round(2)
        df_base.drop(columns=["VALOR", "VALOR_DESC_ORIG", "DIF"], inplace=True)

        emit("Formatando datas...", 55)
        df_aba1 = df_base.copy()
        df_aba1["DATA_LANC"] = _formatar_data(df_aba1["DATA_LANC"])
        df_aba1 = df_aba1[COLUNAS_FINAIS]

        df_aba2 = df_base[df_base["VALOR_DESCONTADO"] > 0].copy()
        df_aba2["DATA_LANC"] = _formatar_data(df_aba2["DATA_LANC"])
        df_aba2 = df_aba2[COLUNAS_FINAIS]

        df_aba3 = df_base[df_base["VALOR_DESCONTADO"] == 0].copy()
        df_aba3["DATA_LANC"] = _formatar_data(df_aba3["DATA_LANC"])
        df_aba3 = df_aba3[COLUNAS_FINAIS]

        emit("Gerando ABA_1...", 65)
        with pd.ExcelWriter(saida, engine="openpyxl") as writer:
            df_aba1.to_excel(writer, index=False, sheet_name="ABA_1")

        emit("Gerando abas DESCONTADOS e NAO_DESCONTADOS...", 78)
        wb = load_workbook(saida)

        emit("Ajustando formatação...", 90)
        _formatar_aba_existente(wb["ABA_1"])
        _ajustar_largura(wb["ABA_1"])
        for nome, df in [("DESCONTADOS", df_aba2), ("NAO_DESCONTADOS", df_aba3)]:
            ws = _escrever_aba(wb, nome, df)
            _ajustar_largura(ws)

        wb.save(saida)

        total = len(df_aba1)
        desc = len(df_aba2)
        n_desc = len(df_aba3)

        emit("done", 100)
        finished_jobs[job_id] = saida
        q.put({
            "done": True,
            "rows": total,
            "descontados": desc,
            "nao_descontados": n_desc,
            "job_id": job_id,
            "file_name": os.path.basename(saida),
        })

    except Exception:
        # Loga o stack trace completo no servidor (logs/app.log), mas devolve
        # uma mensagem genérica e amigável ao usuário.
        logger.exception("Erro ao converter arquivo (job_id=%s)", job_id)
        q.put({
            "error": (
                "Não foi possível converter o arquivo. Verifique se o "
                "layout do .txt está correto e tente novamente."
            )
        })
